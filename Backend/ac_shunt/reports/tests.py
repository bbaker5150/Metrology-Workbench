"""Smoke tests for the Report of Calibration module."""
from io import BytesIO

from django.apps import apps
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIRequestFactory

from . import area_registry, excel, importer, views
from .models import ROCRecord


class ReportsScaffoldTests(TestCase):
    def test_app_is_installed(self):
        self.assertTrue(apps.is_installed("reports"))

    def test_module_info_endpoint(self):
        request = APIRequestFactory().get("/api/reports/info/")
        response = views.module_info(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["module"], "reports")


class ReportsCrudTests(TestCase):
    databases = {"default", "reports"}

    def test_seed_command_is_idempotent(self):
        call_command("seed_rocs")
        self.assertEqual(ROCRecord.objects.count(), 3)
        call_command("seed_rocs")  # re-running must not duplicate rows
        self.assertEqual(ROCRecord.objects.count(), 3)

    def test_area_registry_lists_areas_with_no_seeding(self):
        # Areas are files, not DB rows -- available even on a fresh DB.
        codes = {area["code"] for area in area_registry.list_areas()}
        self.assertEqual(codes, {"AC_SHUNT", "RESISTANCE", "TEMPERATURE"})
        self.assertIsNotNone(area_registry.get_area("AC_SHUNT"))
        self.assertIsNone(area_registry.get_area("NOT_A_REAL_AREA"))

    def test_areas_endpoint_lists_areas(self):
        request = APIRequestFactory().get("/api/reports/areas/")
        response = views.areas(request)
        self.assertEqual(response.status_code, 200)
        codes = {area["code"] for area in response.data}
        self.assertEqual(codes, {"AC_SHUNT", "RESISTANCE", "TEMPERATURE"})

    def test_roc_create_and_delete_round_trip(self):
        request = APIRequestFactory().post("/api/reports/rocs/", {
            "roc_number": "2026-999999",
            "nomenclature": "Test Instrument",
        }, format="json")
        response = views.rocs(request)
        self.assertEqual(response.status_code, 201)
        roc_id = response.data["id"]

        detail_request = APIRequestFactory().delete(f"/api/reports/rocs/{roc_id}/")
        detail_response = views.roc_detail(detail_request, roc_id=roc_id)
        self.assertEqual(detail_response.status_code, 204)
        self.assertFalse(ROCRecord.objects.filter(pk=roc_id).exists())


class ReportsExcelTests(TestCase):
    databases = {"default", "reports"}

    def test_roc_generate_produces_a_valid_workbook(self):
        payload = {
            "roc_number": "2026-000123",
            "nomenclature": "Current Shunt",
            "statements": [{"kind": "technical", "text": "Test statement."}],
            "tables": [{
                "title": "Data",
                "columns": [{"header": "Current", "unit": "(A)"}],
                "rows": [[1.0], [2.0]],
            }],
        }
        request = APIRequestFactory().post("/api/reports/roc/generate/", payload, format="json")
        response = views.roc_generate(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # A corrupt/empty workbook would raise here. Every area shares this
        # one ROC1/ROC2 layout now (see cellmap.TEMPLATE), so no area_code
        # in the payload is needed to pick a template. ROC2 is the real,
        # certificate-style page 2 (build_workbook's own polished grid,
        # matching the PDF preview) -- present because this record has a
        # table; Data Entry is the separate, hand-editable form for
        # round-tripping through Excel Import.
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["ROC1", "ROC2", "Data Entry"])
        self.assertEqual(workbook["ROC1"]["Z52"].value, "Page 1 of 2")
        self.assertEqual(workbook["ROC2"]["A7"].value, "Data")  # table title (not tied)
        # ROC2's data cells are tied to Data Entry (a live formula, not a
        # duplicated literal) -- editing the number in Data Entry updates
        # the certificate's own copy of it in Excel.
        self.assertEqual(workbook["ROC2"]["A10"].value, "=IF('Data Entry'!A62=\"\",\"\",'Data Entry'!A62)")
        self.assertEqual(workbook["Data Entry"]["A62"].value, 1.0)
        # ROC2's table cells aren't a fill-in form -- no yellow input fill.
        self.assertIsNone(workbook["ROC2"]["A10"].fill.patternType)
        # Round-tripping the generated workbook through the importer is a
        # more robust check than pinning exact cell coordinates here.
        parsed = importer.parse_workbook(BytesIO(response.content))
        self.assertEqual(parsed["roc_number"], "2026-000123")
        self.assertEqual(parsed["statements"], [{"kind": "technical", "text": "Test statement."}])
        self.assertEqual(len(parsed["tables"]), 1)
        self.assertEqual(parsed["tables"][0]["columns"][0], {"header": "Current", "unit": "(A)"})
        self.assertEqual(parsed["tables"][0]["rows"][0][0], 1.0)
        self.assertEqual(parsed["tables"][0]["rows"][1][0], 2.0)

    def test_every_area_shares_the_canonical_layout(self):
        # All three areas' ROC Template downloads build the same Data Entry
        # form (cellmap.TEMPLATE); saved-record exports (not tested here)
        # additionally build the shared AC_SHUNT-derived ROC1 certificate.
        for code in ("AC_SHUNT", "RESISTANCE", "TEMPERATURE"):
            request = APIRequestFactory().get("/api/reports/roc/template/", {"area": code})
            response = views.roc_template(request)
            self.assertEqual(response.status_code, 200)
            workbook = load_workbook(BytesIO(response.content))
            self.assertEqual(workbook.sheetnames, ["Data Entry"])
            # The Data Entry sheet's "Measurement Area" field carries the
            # area for round-tripping through Excel Import (see
            # importer.py's parse_workbook).
            self.assertEqual(workbook["Data Entry"]["A5"].value, "Measurement Area")
            self.assertEqual(workbook["Data Entry"]["B5"].value, code)

    def test_roc_template_for_an_unknown_area_returns_404(self):
        request = APIRequestFactory().get("/api/reports/roc/template/", {"area": "NOT_A_REAL_AREA"})
        response = views.roc_template(request)
        self.assertEqual(response.status_code, 404)

    def test_export_preserves_reference_certificate_cell_structure(self):
        reference = load_workbook(excel.TEMPLATES_DIR / "roc_ac_shunt.xlsx")["ROC1"]
        generated = excel.build_workbook({
            "area_code": "AC_SHUNT",
            "nomenclature": "Current Shunt",
            "statements": [{"kind": "technical", "text": "Replacement text."}],
            "tables": [],
        })["ROC1"]
        self.assertEqual(
            set(map(str, generated.merged_cells.ranges)),
            set(map(str, reference.merged_cells.ranges)),
        )
        self.assertEqual(
            {key: dim.width for key, dim in generated.column_dimensions.items()},
            {key: dim.width for key, dim in reference.column_dimensions.items()},
        )
        self.assertEqual(
            {key: dim.height for key, dim in generated.row_dimensions.items()},
            {key: dim.height for key, dim in reference.row_dimensions.items()},
        )
        self.assertEqual(generated.page_setup.orientation, reference.page_setup.orientation)
        self.assertEqual(generated.page_setup.fitToWidth, reference.page_setup.fitToWidth)

    def test_roc_excel_from_saved_record(self):
        call_command("seed_rocs")
        record = ROCRecord.objects.first()
        request = APIRequestFactory().get(f"/api/reports/rocs/{record.pk}/excel/")
        response = views.roc_excel(request, roc_id=record.pk)
        self.assertEqual(response.status_code, 200)
        load_workbook(BytesIO(response.content))  # raises if malformed

    def test_roc_template_for_a_known_area(self):
        request = APIRequestFactory().get("/api/reports/roc/template/", {"area": "AC_SHUNT"})
        response = views.roc_template(request)
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))  # raises if malformed
        # Just the Data Entry form to fill in and upload back through Excel
        # Import -- no certificate page, since there's no record yet to put
        # on one (see excel.py's build_blank_template_workbook).
        self.assertEqual(workbook.sheetnames, ["Data Entry"])
        ws = workbook["Data Entry"]
        # Every field from Manual Input appears as a label/input pair, in
        # ManualInputForm.jsx's order -- area_code leads so a round-tripped
        # upload always carries its area (see importer.py).
        self.assertEqual(ws["A4"].value, "REPORT FIELDS")
        self.assertEqual(ws["A5"].value, "Measurement Area")
        self.assertEqual(ws["B5"].value, "AC_SHUNT")
        self.assertEqual(ws["B5"].fill.fgColor.rgb, "00FFF2CC")
        self.assertEqual(ws["A6"].value, "RoC #")
        self.assertEqual(ws["A7"].value, "Nomenclature")
        self.assertEqual([label for _, label in excel.cellmap.DATA_ENTRY_FIELDS][-1], "Approver Title")
        # Front-page statements and the calibration table(s) both follow.
        self.assertIn("FRONT-PAGE STATEMENTS", [ws.cell(row=r, column=1).value for r in range(1, 30)])
        self.assertIn("CALIBRATION DATA TABLE(S)", [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)])
        # No frozen panes -- the sheet scrolls freely top to bottom.
        self.assertIsNone(ws.freeze_panes)

    def test_roc_parse_round_trips_a_blank_template_download(self):
        # The blank ROC Template has no certificate page (build_
        # blank_template_workbook) -- Excel Import must still accept it and
        # read the area back from the Data Entry sheet alone.
        call_command("seed_rocs")
        generated = views.roc_template(APIRequestFactory().get("/api/reports/roc/template/", {"area": "RESISTANCE"}))
        upload = SimpleUploadedFile(
            "roc_template.xlsx", generated.content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        request = APIRequestFactory().post("/api/reports/roc/parse/", {"file": upload}, format="multipart")
        response = views.roc_parse(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["area_code"], "RESISTANCE")
        self.assertEqual(response.data["nomenclature"], "Resistance Standard")
        # The default calibration table shell has headers but no real data
        # rows yet -- _scan_tables stops at the first blank row.
        self.assertEqual(len(response.data["tables"]), 1)
        self.assertEqual(response.data["tables"][0]["rows"], [])

    def test_roc_parse_round_trips_a_generated_workbook(self):
        call_command("seed_rocs")
        record = ROCRecord.objects.get(area_code="AC_SHUNT")
        generated = views.roc_excel(APIRequestFactory().get(f"/api/reports/rocs/{record.pk}/excel/"), roc_id=record.pk)

        upload = SimpleUploadedFile(
            "roc.xlsx", generated.content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        request = APIRequestFactory().post("/api/reports/roc/parse/", {"file": upload}, format="multipart")
        response = views.roc_parse(request)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["area_code"], "AC_SHUNT")
        self.assertEqual(response.data["manufacturer"], record.manufacturer)
        self.assertEqual(response.data["serial_number"], record.serial_number)
        self.assertEqual(response.data["roc_number"], record.roc_number)
        self.assertEqual(len(response.data["tables"]), len(record.tables))
        self.assertEqual(len(response.data["tables"][0]["rows"]), len(record.tables[0]["rows"]))

    def test_roc_parse_rejects_an_unrecognized_file(self):
        workbook = Workbook()
        workbook.active.title = "Not A ROC"
        buffer = BytesIO()
        workbook.save(buffer)
        upload = SimpleUploadedFile(
            "random.xlsx", buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        request = APIRequestFactory().post("/api/reports/roc/parse/", {"file": upload}, format="multipart")
        response = views.roc_parse(request)
        self.assertEqual(response.status_code, 400)


class ReportsAcShuntPullTests(TestCase):
    databases = {"default", "reports"}

    def test_sessions_list_and_pull_the_seeded_mock_session(self):
        call_command("seed_mock_calibration_session")
        from api.management.commands.seed_mock_calibration_session import MOCK_SESSION_NAME
        from api.models import CalibrationSession

        session = CalibrationSession.objects.get(session_name=MOCK_SESSION_NAME)

        list_request = APIRequestFactory().get("/api/reports/ac-shunt/sessions/")
        list_response = views.ac_shunt_sessions(list_request)
        self.assertEqual(list_response.status_code, 200)
        self.assertTrue(list_response.data["available"])
        self.assertTrue(any(s["id"] == session.pk for s in list_response.data["sessions"]))

        pull_request = APIRequestFactory().get(f"/api/reports/ac-shunt/sessions/{session.pk}/pull/")
        pull_response = views.ac_shunt_session_pull(pull_request, session_id=session.pk)
        self.assertEqual(pull_response.status_code, 200)
        self.assertEqual(pull_response.data["serial_number"], "MOCK-UUT")
        self.assertEqual(len(pull_response.data["tables"]), 1)
        self.assertEqual(len(pull_response.data["tables"][0]["rows"]), 10)  # 5 pairs x Forward/Reverse

    def test_pull_missing_session_returns_404(self):
        request = APIRequestFactory().get("/api/reports/ac-shunt/sessions/999999/pull/")
        response = views.ac_shunt_session_pull(request, session_id=999999)
        self.assertEqual(response.status_code, 404)
