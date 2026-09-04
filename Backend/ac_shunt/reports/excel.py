"""
Excel (.xlsx) generation for the Report of Calibration module.

Writes into a real, trimmed NPSL workbook template per measurement area
(``reports/templates/*.xlsx`` — see ``cellmap.py`` for exactly which real
report each one traces to and which cell each field maps to), rather than
building a page from scratch: every fixed-position field (letterhead,
instrument identity, statement paragraphs, environment, signatures, footer)
lands in the exact cell it occupies in a real completed ROC — the fonts,
merges, and column widths are literally the source file's — so the output
is the same page, not a same-content reconstruction of it.

Only the measurement data table(s) are generated fresh per record,
immediately below the page-2 header: table shape (column count, row count)
varies per instrument and can't be pinned to fixed template cells the way
the fields above can. They're written using the templates' own technique —
a dense ~55-column grid with logical columns spanning several real columns
via merges — so a generated table still visually matches its neighbors.

``importer.py`` reads a filled-in copy of these same templates back using
the same cellmap, so generation and parsing can't drift out of sync.
"""
from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import cellmap

TEMPLATES_DIR = Path(__file__).resolve().parent / "templates"

DATE_FIELDS = {"calibration_date", "due_date", "issue_date"}
_MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]


def _format_date(value):
    """ISO ('2026-03-02') -> NPSL's own DDMMMYYYY style ('02MAR2026'), same
    as CalibrationPDF.jsx's formatDate. Anything else passes through as-is
    rather than erroring, e.g. free text a user already typed by hand."""
    if not value:
        return value
    try:
        parsed = date.fromisoformat(str(value)[:10])
    except ValueError:
        return value
    return f"{parsed.day:02d}{_MONTHS[parsed.month - 1]}{parsed.year}"

BODY_FONT = Font(name="Times New Roman", size=12)
BOLD_BODY_FONT = Font(name="Times New Roman", size=12, bold=True)
HEADER_FONT = Font(name="Times New Roman", size=11, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
JUSTIFY = Alignment(horizontal="justify", vertical="top", wrap_text=True)
THIN = Side(style="thin")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
INPUT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
INSTRUCTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")

# The real templates' own working width (columns A:BC) -- reused for
# generated tables so they sit on the same grid as the surrounding page.
GRID_COLUMNS = 55

# Front-page statement kinds in NPSL's canonical order (matches
# ManualInputForm.jsx's STATEMENT_LABELS and CalibrationPDF.jsx's
# technical/statements_rest split). A kind with no cell mapping for a given
# area (see cellmap.py) is simply not written.
STATEMENT_ORDER = ["technical", "results_location", "special", "uncertainty", "traceability", "reproduction"]


def _set(ws, cell_ref, value):
    if cell_ref and value not in (None, ""):
        ws[cell_ref] = value


def _tie_formula(data_entry_cell_ref):
    """A live reference to a Data Entry cell, blank instead of 0 when that
    cell is empty -- see build_workbook's REPORT FIELDS / FRONT-PAGE
    STATEMENTS loops and its calibration-table tie."""
    ref = f"'Data Entry'!{data_entry_cell_ref}"
    return f'=IF({ref}="","",{ref})'


def _tie_inline_result_formula(data_entry_row):
    """A live reference to one Data Entry inline-results row (columns
    A-D: Label, Value, Label 2, Value 2), reconstructing the same
    "label = value     label2 = value2" text build_workbook writes literally
    when there's no Data Entry sheet to tie to -- each pair only contributes
    if both its label and value are non-blank, same as the literal version."""
    a, b, c, d = (f"'Data Entry'!{col}{data_entry_row}" for col in "ABCD")
    pair1 = f'IF(AND({a}<>"",{b}<>""),{a}&" = "&{b},"")'
    pair2 = f'IF(AND({c}<>"",{d}<>""),{c}&" = "&{d},"")'
    return f'=TEXTJOIN("     ",TRUE,{pair1},{pair2})'


def _merge_for(ws, cell_ref):
    """The merged range anchored at cell_ref, if any."""
    cell = ws[cell_ref]
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row == cell.row and merged_range.min_col == cell.column:
            return merged_range
    return None


def _apply_border(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = CELL_BORDER


def _column_spans(n):
    """Evenly split GRID_COLUMNS into n logical (start, end) column spans,
    1-indexed and inclusive, the remainder going to the last column."""
    span = GRID_COLUMNS // max(n, 1)
    spans = [(1 + i * span, 1 + (i + 1) * span - 1) for i in range(n)]
    if spans:
        spans[-1] = (spans[-1][0], GRID_COLUMNS)
    return spans


def _write_paragraph(ws, row, text):
    if not text:
        return row
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = BODY_FONT
    cell.alignment = JUSTIFY
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=GRID_COLUMNS)
    lines = max(1, -(-len(text) // 145))
    ws.row_dimensions[row].height = 15.6 * lines
    return row + 2


def _write_table(ws, row, table, *, input_mode=False, tie_layout=None):
    """`tie_layout` (from _build_data_entry_sheet's per-table return value:
    {"data_start_row", "width"}), when given, ties each data cell to its
    Data Entry counterpart instead of writing the value literally -- editing
    the number in Data Entry updates the certificate page's own copy of it
    in Excel. Only real columns (0..width-1) are tied; Data Entry's extra
    "Additional N" growth columns have no certificate-page counterpart to
    tie to."""
    row = _write_paragraph(ws, row, table.get("intro_text"))
    if table.get("title"):
        cell = ws.cell(row=row, column=1, value=table["title"])
        cell.font = BOLD_BODY_FONT
        cell.alignment = CENTER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=GRID_COLUMNS)
        row += 2

    columns = table.get("columns") or []
    if not columns:
        return row + 1
    spans = _column_spans(len(columns))

    for (start, end), column in zip(spans, columns):
        header_text = "\n".join(filter(None, [column.get("header"), column.get("unit")]))
        cell = ws.cell(row=row, column=start, value=header_text)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        if end > start:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        _apply_border(ws, row, start, end)
    ws.row_dimensions[row].height = 30
    row += 1

    for row_index, data_row in enumerate(table.get("rows") or []):
        for col_index, ((start, end), value) in enumerate(zip(spans, data_row)):
            if tie_layout is not None and col_index < tie_layout["width"]:
                de_ref = f"{get_column_letter(col_index + 1)}{tie_layout['data_start_row'] + row_index}"
                cell_value = _tie_formula(de_ref)
            else:
                cell_value = value if value not in (None,) else ""
            cell = ws.cell(row=row, column=start, value=cell_value)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            if input_mode:
                cell.fill = INPUT_FILL
            if end > start:
                ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            _apply_border(ws, row, start, end)
        row += 1

    return row + 2


def _fit_to_one_page_wide(ws):
    """Force print-scale-to-fit regardless of the template's own column
    widths, and open in Page Layout view (matching the real AC_SHUNT source
    file) rather than Normal view. Excel -- and any thumbnailer/previewer
    that takes its cue from the saved view mode -- renders Page Layout view
    already reflowed to the page; Normal view shows the raw, unscaled cell
    grid and lets far-right content run past the printable page instead of
    scaling down to fit, which is what a from-scratch openpyxl Workbook()
    (the RESISTANCE/TEMPERATURE templates' starting point, unlike AC_SHUNT's
    which is a trimmed copy of a real, already page-authored file) defaults
    to."""
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.view = "pageLayout"


def _make_calibration_data_sheet(wb, area):
    """Create the second worksheet used when a source ROC stored its result
    pages below page 1 instead of on a separate tab (the temperature ROC).
    It uses the same 55-column grid and header positions as the two-tab lab
    ROCs, keeping the first worksheet untouched and source-formatted."""
    ws = wb.create_sheet(area["page2_sheet"])
    ws.sheet_view.showGridLines = False
    ws.sheet_view.view = "pageLayout"
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    for column in range(1, GRID_COLUMNS + 1):
        ws.column_dimensions[ws.cell(1, column).column_letter].width = 2.35

    labels = {
        "A1": "Nomenclature:", "AM1": "Calibration Date:",
        "A2": "Manufacturer:", "AM2": "Due Date:",
        "A3": "Model:", "A4": "Serial:", "A45": "RoC #:",
    }
    for ref, value in labels.items():
        ws[ref] = value
        ws[ref].font = BODY_FONT
    for cell_range in ("A1:J1", "K1:AL1", "AM1:AV1", "AW1:BC1",
                       "A2:J2", "K2:AL2", "AM2:AV2", "AW2:BC2",
                       "A3:J3", "K3:BC3", "A4:J4", "K4:BC4",
                       "A45:D45", "E45:BC45", "A46:BC46"):
        ws.merge_cells(cell_range)
    _fit_to_one_page_wide(ws)
    return ws


def build_workbook(data: dict, *, template_mode=False, with_data_entry_tie=False):
    """Build a Workbook from a ROC payload dict (the same shape ManualInputForm
    edits and ROCRecordSerializer produces).

    `with_data_entry_tie=True` (build_template_workbook's downloads) also
    builds the Data Entry sheet (_build_data_entry_sheet) and ties every
    certificate-page value to it -- REPORT FIELDS / FRONT-PAGE STATEMENTS
    (fixed cells, always tied), plus Inline Results and each calibration
    table's data cells (their Data Entry row/column only known once Data
    Entry itself has been built, hence building it first, here, rather than
    as a separate step afterward). `False` (the default; every other
    caller, including the tests that inspect build_workbook's output
    directly) writes plain literal values and no Data Entry sheet at all."""
    area = cellmap.TEMPLATE
    wb = load_workbook(TEMPLATES_DIR / area["template"])
    ws1 = wb[area["page1_sheet"]]
    # The trimmed temperature source contains legacy result-page formatting
    # below the certificate face. Limit printing to the source-formatted
    # first page; current raw data belongs on the dedicated second worksheet.
    if area["page1_sheet"] == "Report of Calibration":
        for merged_range in list(ws1.merged_cells.ranges):
            if merged_range.min_row >= 62:
                ws1.unmerge_cells(str(merged_range))
        if ws1.max_row > 61:
            ws1.delete_rows(62, ws1.max_row - 61)
        ws1.print_area = "A1:AZ61"
    # Do not normalize, resize, or otherwise restyle the certificate page.
    # Its merges, widths, row heights, print settings, and page view come
    # directly from the lab ROC template and are the formatting contract.
    if area["page2_sheet"] not in wb.sheetnames:
        _make_calibration_data_sheet(wb, area)
    else:
        _fit_to_one_page_wide(wb[area["page2_sheet"]])

    # Built before anything below ties to it, so each calibration table's
    # Data Entry row/column is already known (_build_data_entry_sheet
    # returns it) by the time the ws2 loop needs it.
    table_layouts = None
    if with_data_entry_tie:
        _, table_layouts = _build_data_entry_sheet(wb, data)

    # REPORT FIELDS / FRONT-PAGE STATEMENTS are *tied*: the certificate cell
    # holds a live formula pointing at its Data Entry input cell (see
    # cellmap.data_entry_field_row/_statement_row) instead of a duplicated
    # value, so editing either sheet in Excel updates the other. Inline
    # Results and the calibration table(s) are tied the same way -- see
    # below -- once a Data Entry sheet exists to tie to (with_data_entry_tie).
    fields = area["fields"]
    for name, cell_ref in fields.items():
        de_row = cellmap.data_entry_field_row(name)
        if de_row is not None and with_data_entry_tie:
            ws1[cell_ref] = _tie_formula(f"B{de_row}")
        else:
            value = data.get(name)
            _set(ws1, cell_ref, _format_date(value) if name in DATE_FIELDS else value)

    for kind, cell_ref in area["statements"].items():
        de_row = cellmap.data_entry_statement_row(kind)
        if de_row is not None and with_data_entry_tie:
            ws1[cell_ref] = _tie_formula(f"B{de_row}")

    tables = data.get("tables") or []
    pages = 2 if tables else 1
    if "page_label" in fields:
        _set(ws1, fields["page_label"], f"Page 1 of {pages}")

    same_sheet = area["page1_sheet"] == area["page2_sheet"]
    if data.get("inline_results"):
        # Sits right below the technical statement, above the environment
        # block -- placed generically since no real template has a fixed
        # cell for an arbitrary list of front-page coefficients (see
        # module docstring on tables for the same reasoning).
        anchor_cell = area["statements"].get("technical")
        if anchor_cell:
            technical_merge = _merge_for(ws1, anchor_cell)
            last_row = technical_merge.max_row if technical_merge else ws1[anchor_cell].row
            row = last_row + 2
        else:
            row = 20
        for index, line in enumerate(data["inline_results"]):
            if with_data_entry_tie:
                de_row = cellmap.DATA_ENTRY_INLINE_RESULTS_DATA_START_ROW + index
                cell_value = _tie_inline_result_formula(de_row)
            else:
                cells = [str(c) for c in line if c not in (None, "")]
                if not cells:
                    continue
                pairs = [f"{cells[i]} = {cells[i + 1]}" for i in range(0, len(cells) - 1, 2)]
                cell_value = "     ".join(pairs)
            cell = ws1.cell(row=row, column=1, value=cell_value)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=GRID_COLUMNS)
            row += 1

    if not tables:
        if not same_sheet:
            del wb[area["page2_sheet"]]
        return wb

    ws2 = wb[area["page2_sheet"]]
    for name, cell_ref in area["page2_fields"].items():
        if name == "page_label":
            _set(ws2, cell_ref, f"Page 2 of {pages}")
        else:
            value = data.get(name)
            _set(ws2, cell_ref, _format_date(value) if name in DATE_FIELDS else value)

    # The real lab-original source (roc_ac_shunt.xlsx) has its own stray
    # border formatting scattered below the header, left over from that
    # specific completed report's own table -- empty cells with a border and
    # no value, at positions unrelated to whatever shape *this* record's
    # table(s) turn out to be. Left alone, they show up as random bordered
    # boxes wherever our own table doesn't happen to cover them. Clear
    # everything from table_start_row down before writing our own table(s)
    # so nothing but our own formatting shows.
    for r in range(area["table_start_row"], max(ws2.max_row, area["table_start_row"]) + 1):
        for c in range(1, GRID_COLUMNS + 1):
            cell = ws2.cell(row=r, column=c)
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)

    row = area["table_start_row"]
    for index, table in enumerate(tables):
        layout = table_layouts[index] if table_layouts else None
        row = _write_table(ws2, row, table, input_mode=template_mode, tie_layout=layout)

    if template_mode:
        # Title, spacer, and header occupy the first three table rows.
        ws2.freeze_panes = ws2.cell(row=area["table_start_row"] + 3, column=1)
        note = ws2.cell(row=area["table_start_row"] - 1, column=1)
        note.value = "Enter calibration raw data in the yellow cells. Rename headers or add rows as needed."
        note.font = Font(name="Times New Roman", size=10, italic=True)
        note.fill = INSTRUCTION_FILL
        note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws2.merge_cells(start_row=area["table_start_row"] - 1, start_column=1,
                        end_row=area["table_start_row"] - 1, end_column=GRID_COLUMNS)

    return wb


def _write_section_header(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Times New Roman", size=12, bold=True, underline="single")
    return row + 1


def _write_form_field_row(ws, row, label, value):
    """One "Field label | input cell" row of the Data Entry form -- the
    left/right pairing the app's whole Manual Input tab is built from."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = BODY_FONT
    label_cell.alignment = Alignment(vertical="center")
    value_cell = ws.cell(row=row, column=2, value=value if value not in (None, "") else "")
    value_cell.font = BODY_FONT
    value_cell.alignment = Alignment(vertical="center", wrap_text=True)
    value_cell.fill = INPUT_FILL
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    _apply_border(ws, row, 1, 5)
    return row + 1


def _write_form_statement_row(ws, row, label, text):
    """Same field/value pairing as _write_form_field_row, but the value
    cell spans STATEMENT_ROW_HEIGHT rows and wraps like a paragraph, for a
    front-page statement's prose."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = BODY_FONT
    label_cell.alignment = Alignment(vertical="top")
    end_row = row + cellmap.STATEMENT_ROW_HEIGHT - 1
    ws.merge_cells(start_row=row, start_column=2, end_row=end_row, end_column=8)
    value_cell = ws.cell(row=row, column=2, value=text or "")
    value_cell.font = BODY_FONT
    value_cell.alignment = JUSTIFY
    value_cell.fill = INPUT_FILL
    for r in range(row, end_row + 1):
        _apply_border(ws, r, 1, 8)
        ws.row_dimensions[r].height = 20
    return end_row + 1


def _write_inline_results_table(ws, row, inline_results, *, min_rows=8):
    for col, header in enumerate(cellmap.INLINE_RESULT_HEADERS, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = INPUT_FILL
        cell.border = CELL_BORDER
    row += 1
    count = max(len(inline_results), min_rows)
    for i in range(count):
        values = (list(inline_results[i]) if i < len(inline_results) else []) + ["", "", "", ""]
        for col, value in enumerate(values[:4], start=1):
            cell = ws.cell(row=row, column=col, value=value if value not in (None,) else "")
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.fill = INPUT_FILL
            cell.border = CELL_BORDER
        row += 1
    return row


def _write_calibration_table(ws, row, table, *, area_code="", extra_columns=5,
                              min_extra_rows=10, min_total_rows=25):
    """One calibration data table: title (+ optional intro paragraph),
    header row, then data rows padded well past whatever real data is
    present, plus a handful of pre-formatted extra columns -- so there's
    always a lot of blank, ready-to-use rows and columns to grow into
    without having to fight Excel merges to add them. Every cell is a plain,
    unmerged cell (unlike the certificate page's grid), so inserting more of
    either is trivial too."""
    columns = table.get("columns") or [{"header": h, "unit": u} for h, u in
                                        (RAW_DATA_COLUMNS.get(area_code) or DEFAULT_RAW_DATA_COLUMNS)]
    data_rows = table.get("rows") or []
    width = len(columns)
    total_width = width + extra_columns
    title_span = max(total_width, 12)

    if table.get("intro_text"):
        cell = ws.cell(row=row, column=1, value=table["intro_text"])
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=title_span)
        row += 1

    title_cell = ws.cell(row=row, column=1, value=table.get("title") or "Calibration Raw Data")
    title_cell.font = BOLD_BODY_FONT
    title_cell.alignment = CENTER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=title_span)
    row += 1

    for col in range(1, total_width + 1):
        if col <= width:
            definition = columns[col - 1]
            header_text = "\n".join(filter(None, [definition.get("header"), definition.get("unit")]))
        else:
            header_text = f"Additional {col - width}"
        cell = ws.cell(row=row, column=col, value=header_text)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = INPUT_FILL
        cell.border = CELL_BORDER
    row += 1
    data_start_row = row

    height = max(len(data_rows) + min_extra_rows, min_total_rows)
    for r in range(height):
        values = data_rows[r] if r < len(data_rows) else []
        for col in range(1, total_width + 1):
            value = values[col - 1] if col <= len(values) else ""
            cell = ws.cell(row=row, column=col, value=value if value not in (None,) else "")
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.fill = INPUT_FILL
            cell.border = CELL_BORDER
        row += 1

    layout = {"data_start_row": data_start_row, "width": width}
    return row + 1, layout  # blank spacer row before the next table


def _build_data_entry_sheet(wb, data):
    """The synthetic "Data Entry" page 2: every Manual Input field as a
    left-label/right-input pair, front-page statements, the optional inline
    coefficients, and the calibration data table(s) at the bottom. Built
    from `data`, so this is identical whether `data` is empty (a blank ROC
    Template download) or a saved/drafted record's full payload (a
    prefilled download) -- see build_template_workbook.

    Returns (ws, table_layouts): table_layouts is one
    {"data_start_row", "width"} dict per table in `data["tables"]`, in
    order -- used by build_workbook (with_data_entry_tie=True) to tie the
    certificate page's own copy of each calibration table to these same
    cells."""
    ws = wb.create_sheet("Data Entry")
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJKLMNOPQR":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "ROC Data Entry Form"
    ws["A1"].font = Font(name="Times New Roman", size=16, bold=True)
    ws["A2"] = ("Enter or edit values in the yellow cells below, then upload this workbook through "
                "Excel Import to load it into the app. Add rows or columns to the calibration "
                "data table(s) at the bottom as needed.")
    ws["A2"].font = Font(name="Times New Roman", size=10, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 28

    row = cellmap.DATA_ENTRY_FIELDS_HEADER_ROW
    row = _write_section_header(ws, row, cellmap.DATA_ENTRY_SECTION_HEADERS["fields"])
    for key, label in cellmap.DATA_ENTRY_FIELDS:
        value = data.get(key)
        if key in DATE_FIELDS:
            value = _format_date(value)
        row = _write_form_field_row(ws, row, label, value)

    row += 1
    row = _write_section_header(ws, row, cellmap.DATA_ENTRY_SECTION_HEADERS["statements"])
    statements_by_kind = {s.get("kind"): s.get("text", "") for s in (data.get("statements") or [])}
    for kind, label in cellmap.DATA_ENTRY_STATEMENTS:
        row = _write_form_statement_row(ws, row, label, statements_by_kind.get(kind, ""))

    # From the fixed cellmap row (not the natural cursor) -- guarantees this
    # can't silently drift from DATA_ENTRY_INLINE_RESULTS_DATA_START_ROW,
    # which build_workbook's inline-results tie relies on.
    row = cellmap.DATA_ENTRY_INLINE_RESULTS_HEADER_ROW
    row = _write_section_header(ws, row, cellmap.DATA_ENTRY_SECTION_HEADERS["inline_results"])
    note = ws.cell(row=row, column=1, value="One row per line: Label | Value | Label 2 | Value 2.")
    note.font = Font(name="Times New Roman", size=10, italic=True)
    row += 1
    row = _write_inline_results_table(ws, row, data.get("inline_results") or [])
    assert row == cellmap.DATA_ENTRY_INLINE_RESULTS_DATA_START_ROW + max(len(data.get("inline_results") or []), 8), \
        "DATA_ENTRY_INLINE_RESULTS_DATA_START_ROW drifted from _write_inline_results_table's actual layout"

    row += 1
    row = _write_section_header(ws, row, cellmap.DATA_ENTRY_SECTION_HEADERS["calibration"])
    note = ws.cell(row=row, column=1,
                    value="Enter calibration raw data in the yellow cells. Rename headers, or add rows/columns, as needed.")
    note.font = Font(name="Times New Roman", size=10, italic=True)
    note.fill = INSTRUCTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    row += 2  # instruction row, then a blank spacer row before the first table

    tables = data.get("tables") or [{}]
    area_code = data.get("area_code") or ""
    table_layouts = []
    for table in tables:
        row, layout = _write_calibration_table(ws, row, table, area_code=area_code)
        table_layouts.append(layout)

    last_col = max(14, max((len(t.get("columns") or []) + 5 for t in tables), default=14))
    ws.print_area = f"A1:{ws.cell(row=1, column=last_col).column_letter}{row}"
    return ws, table_layouts


def build_template_workbook(data: dict):
    """Build a filled-in download: the certificate page from the lab
    template, its real page-2 calibration data page when there are tables
    (both via build_workbook -- the same polished, read-only look the app's
    PDF preview renders, not a fill-in form), plus the synthetic Data Entry
    page (_build_data_entry_sheet). Every value on the certificate and its
    real page 2 -- REPORT FIELDS, FRONT-PAGE STATEMENTS, Inline Results, and
    each calibration table's data cells -- is a live formula tied to Data
    Entry (build_workbook's with_data_entry_tie=True): edit either page in
    Excel and the other updates too. Used for both download buttons that
    have a real record behind them (draft generate, saved-record export) --
    see build_download_workbook. The blank "ROC Template" download has no
    record yet, so it skips the certificate page entirely instead of
    shipping an empty one -- see build_blank_template_workbook.
    """
    return build_workbook(data, with_data_entry_tie=True)


def build_download_workbook(data: dict):
    """Public builder used by both download buttons that have a real record
    behind them (draft generate and saved-record export) -- see
    build_template_workbook."""
    return build_template_workbook(data)


def build_blank_template_workbook(data: dict):
    """Build the blank "ROC Template" download (views.roc_template): the
    Data Entry form page alone, with no certificate page. There's no real
    record yet at this point -- just an area's defaults -- so a mostly-empty
    certificate page (its REPORT FIELDS / FRONT-PAGE STATEMENTS cells are
    live formulas pointing at Data Entry anyway, see build_workbook) has
    nothing useful to show; the Data Entry form is the actual thing to fill
    out and upload back through Excel Import."""
    wb = Workbook()
    wb.remove(wb.active)  # openpyxl always creates one default empty sheet
    _build_data_entry_sheet(wb, data)  # table_layouts unused -- no certificate page here to tie
    return wb


def workbook_to_bytes(workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# Default calibration-table columns per area, used by
# _write_calibration_table when a record has no table of its own yet (a
# blank ROC Template download) -- gives the user real, area-appropriate
# headers to start from instead of a generic "Column 1..N". Renaming,
# adding, or removing any of them is fine: importer.py's table scan is
# column/row-count agnostic, so it comes back in correctly either way.
RAW_DATA_COLUMNS = {
    "AC_SHUNT": [
        ("Current", "(A)"), ("Frequency", "(Hz)"), ("Direction", ""),
        ("ΔUUT", "(ppm)"), ("Expanded Unc.", "(ppm)"), ("k", ""),
    ],
    "RESISTANCE": [
        ("Calibration Point", ""), ("Nominal Resistance", "(Ω)"),
        ("Measured Resistance", "(Ω)"), ("Correction", "(ppm)"),
        ("Expanded Unc.", "(ppm)"), ("k", ""),
    ],
    "TEMPERATURE": [
        ("Calibration Point", "(°C)"), ("Standard Reading", "(Ω)"),
        ("UUT Reading", "(Ω)"), ("Correction", "(°C)"),
        ("Expanded Unc.", "(°C)"), ("k", ""),
    ],
}

DEFAULT_RAW_DATA_COLUMNS = [
    ("Calibration Point", ""), ("Standard Reading", ""),
    ("UUT Reading", ""), ("Correction", ""),
    ("Expanded Unc.", ""), ("k", ""),
]
