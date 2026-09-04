"""
Parse an uploaded ROC workbook (a filled-in copy of this module's one
canonical template, or a workbook hand-conformed to its layout) back into
the same payload shape ``excel.py`` writes from and ``ROCRecordSerializer``
speaks.

Uses the exact same ``cellmap.py`` as the writer for every fixed field, so
the two directions can't drift out of sync — a cell that moves in
``cellmap.py`` moves for both reading and writing at once. Since every area
now shares one layout (``cellmap.TEMPLATE``), which *area* a workbook
belongs to is no longer needed to know where to read cells from -- it's
just metadata, recovered from the Data Entry sheet's "Measurement Area"
field when present (see ``area_code`` in ``cellmap.DATA_ENTRY_FIELDS``),
else from the caller-supplied ``area_hint`` (the frontend's Excel Import
area dropdown), else left blank.

Whenever a "Data Entry" sheet is present -- every one of this app's own
downloads has one, blank template or prefilled -- ``parse_workbook`` reads
*only* that sheet and ignores the certificate page entirely, even a
prefilled one with real values on it: the certificate page's REPORT
FIELDS / FRONT-PAGE STATEMENTS cells are live formulas pointing right back
at the Data Entry sheet (see excel.py's ``build_workbook``), so Data Entry
is the one place a value is guaranteed to be real rather than a formula
result. Only a workbook with no Data Entry sheet at all -- a genuine lab-
original, or one hand-conformed to the certificate layout -- falls back to
reading the certificate page and its own page-2 data table directly.

The measurement-data table(s) can't be read that way (see cellmap.py's
``table_start_row`` docstring: shape varies per instrument), so
``_scan_tables`` figures the shape out from the sheet itself: it walks down
from ``table_start_row``, and for each table it finds, reads the header
row's *populated* cells (merged or not) to learn the actual column
boundaries, then reads data rows using those same boundaries until a blank
row ends the table. Column count, column width, and row count are all
whatever the sheet actually has -- nothing here is pinned to a fixed table
shape, so a user can rename headers, add/remove columns, or add/remove rows
freely and it still comes back correctly.
"""
from datetime import date, datetime

from openpyxl import load_workbook

from . import area_registry, cellmap

_MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
MAX_SCAN_COLUMN = 60
DATE_FIELDS = {"calibration_date", "due_date", "issue_date"}


class UnsupportedWorkbook(Exception):
    pass


def _parse_date(value):
    """NPSL's DDMMMYYYY ('02MAR2026') or an Excel datetime -> ISO, the
    reverse of excel.py's _format_date. Anything else passes through as
    typed (free text a lab metrologist entered by hand)."""
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    if len(text) == 9 and text[:2].isdigit() and text[2:5].upper() in _MONTHS and text[5:].isdigit():
        month = _MONTHS.index(text[2:5].upper()) + 1
        return date(int(text[5:]), month, int(text[:2])).isoformat()
    return text


def _row_populated_columns(ws, row, max_col=MAX_SCAN_COLUMN):
    """[(start_col, end_col), ...] for each populated cell in this row, one
    span per cell -- merged cells contribute their full merge width, plain
    cells contribute just themselves. This is the actual column layout of
    whatever table the sheet has, not an assumption about it."""
    merges_here = {
        m.min_col: m.max_col
        for m in ws.merged_cells.ranges
        if m.min_row <= row <= m.max_row and m.min_col <= max_col
    }
    spans = []
    col = 1
    while col <= max_col:
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            end = merges_here.get(col, col)
            spans.append((col, end))
            col = end + 1
        else:
            col += 1
    return spans


def _row_is_blank(ws, row, max_col=MAX_SCAN_COLUMN):
    return not any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, max_col + 1))


def _split_header(text):
    lines = str(text or "").split("\n")
    return {"header": lines[0] if lines else "", "unit": lines[1] if len(lines) > 1 else ""}


def _scan_tables(ws, start_row, max_row=None, max_col=MAX_SCAN_COLUMN):
    """Walk the data-table region starting at start_row, returning every
    table found as {title, intro_text, columns, rows}. See module docstring
    for the shape-detection approach."""
    max_row = max_row or ws.max_row
    tables = []
    row = start_row

    def skip_blank():
        nonlocal row
        while row <= max_row and _row_is_blank(ws, row, max_col):
            row += 1

    skip_blank()
    while row <= max_row:
        first_cell = str(ws.cell(row=row, column=1).value or "").strip().lower()
        if first_cell.startswith("roc #") or first_cell.startswith("page "):
            break  # reached the page footer, not another table

        title = ""
        intro_text = ""
        # A row with exactly one populated cell, in column 1, spanning the
        # row (a merged full-width row like the writer produces) is prose,
        # not a header -- collect up to two such rows (intro paragraph,
        # then title) before looking for the real header row.
        for _ in range(2):
            spans = _row_populated_columns(ws, row, max_col)
            if len(spans) == 1 and spans[0][0] == 1 and spans[0][1] > 10:
                text = str(ws.cell(row=row, column=1).value)
                if len(text) > 60 and not intro_text:
                    intro_text = text
                else:
                    title = text
                row += 1
                skip_blank()
            else:
                break

        header_spans = _row_populated_columns(ws, row, max_col)
        if not header_spans:
            break
        columns = [_split_header(ws.cell(row=row, column=start).value) for start, _ in header_spans]
        row += 1

        rows = []
        while row <= max_row and not _row_is_blank(ws, row, max_col):
            rows.append([ws.cell(row=row, column=start).value for start, _ in header_spans])
            row += 1

        tables.append({"title": title, "intro_text": intro_text, "columns": columns, "rows": rows})
        skip_blank()

    return tables


def _find_marker_row(ws, label, start=1, max_row=None):
    """The row whose column-A value is exactly `label` -- used to anchor
    each Data Entry form section (see excel.py's DATA_ENTRY_SECTION_HEADERS)
    without pinning to a fixed row number, so unrelated content above a
    section growing or shrinking can't misalign it."""
    max_row = max_row or ws.max_row
    for row in range(start, max_row + 1):
        if str(ws.cell(row=row, column=1).value or "").strip() == label:
            return row
    return None


def _parse_data_entry_sheet(ws):
    """Read the synthetic "Data Entry" form sheet (excel.py's
    _build_data_entry_sheet) back into field/statement overrides, inline
    results, and calibration tables. Walks cellmap.DATA_ENTRY_FIELDS /
    DATA_ENTRY_STATEMENTS in the same order the writer used, each section
    anchored off its marker row; the calibration table(s) are shape-agnostic
    like _scan_tables above, since a user can freely add or remove rows and
    columns there."""
    headers = cellmap.DATA_ENTRY_SECTION_HEADERS
    fields, statements, inline_results, tables = {}, [], [], []

    fields_row = _find_marker_row(ws, headers["fields"])
    if fields_row:
        row = fields_row + 1
        for key, _label in cellmap.DATA_ENTRY_FIELDS:
            value = ws.cell(row=row, column=2).value
            if key in DATE_FIELDS:
                value = _parse_date(value)
            if value not in (None, ""):
                fields[key] = value
            row += 1

    statements_row = _find_marker_row(ws, headers["statements"], start=fields_row or 1)
    if statements_row:
        row = statements_row + 1
        for kind, _label in cellmap.DATA_ENTRY_STATEMENTS:
            text = ws.cell(row=row, column=2).value
            if text not in (None, ""):
                statements.append({"kind": kind, "text": str(text)})
            row += cellmap.STATEMENT_ROW_HEIGHT

    inline_row = _find_marker_row(ws, headers["inline_results"], start=statements_row or 1)
    if inline_row:
        row = inline_row + 3  # marker row, instruction line, column-header row
        while row <= ws.max_row and not _row_is_blank(ws, row, max_col=4):
            values = [ws.cell(row=row, column=c).value for c in range(1, 5)]
            if any(value not in (None, "") for value in values):
                inline_results.append(["" if value is None else value for value in values])
            row += 1

    calibration_row = _find_marker_row(ws, headers["calibration"], start=inline_row or 1)
    if calibration_row:
        row = calibration_row + 2  # marker row, then the instruction note row
        tables = _scan_tables(ws, row)

    return {"fields": fields, "statements": statements, "inline_results": inline_results, "tables": tables}


def parse_workbook(file_obj, area_hint=None):
    """file_obj: a file-like object (as from Django's request.FILES).
    Returns a ROC payload dict, or raises UnsupportedWorkbook."""
    try:
        workbook = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user as-is
        raise UnsupportedWorkbook(
            "Couldn't open this file as an Excel workbook (.xlsx/.xlsm only -- "
            "an old .xls export needs to be re-saved as .xlsx first)."
        ) from exc

    area = cellmap.TEMPLATE
    data_entry_ws = workbook["Data Entry"] if "Data Entry" in workbook.sheetnames else None
    if area["page1_sheet"] not in workbook.sheetnames and data_entry_ws is None:
        raise UnsupportedWorkbook(
            f"This workbook doesn't look like a ROC template (expected a "
            f"'{area['page1_sheet']}' or 'Data Entry' sheet)."
        )

    area_code = area_hint or ""
    data = {"area_code": area_code, "area_name": area_code.replace("_", " ").title() if area_code else ""}

    if data_entry_ws is not None:
        # One of the app's own downloads (blank template, or a prefilled
        # draft/saved-record export with both pages -- see excel.py's
        # build_blank_template_workbook / build_template_workbook). Read
        # *only* the Data Entry form: it's the one place every field,
        # statement, inline result, and calibration table is guaranteed to
        # be a real, current value -- the certificate page's own copies are
        # live formulas pointing right back at these same Data Entry cells
        # (see build_workbook), so re-reading them here would just be
        # reading this same data a second, formula-shaped way.
        form = _parse_data_entry_sheet(data_entry_ws)
        data.update(form["fields"])
        data["statements"] = form["statements"]
        data["inline_results"] = form["inline_results"]
        data["tables"] = form["tables"]
    else:
        # A real lab-original workbook (or one hand-conformed to the
        # certificate layout) with no Data Entry sheet of its own -- read
        # the certificate page and its real page-2 data table(s) directly.
        ws1 = workbook[area["page1_sheet"]]
        ws2 = workbook[area["page2_sheet"]] if area["page2_sheet"] in workbook.sheetnames else None
        for name, cell_ref in area["fields"].items():
            value = ws1[cell_ref].value
            if name in DATE_FIELDS:
                value = _parse_date(value)
            data[name] = "" if value is None else value

        statements = []
        for kind, cell_ref in area["statements"].items():
            text = ws1[cell_ref].value
            if text not in (None, ""):
                statements.append({"kind": kind, "text": str(text)})
        data["statements"] = statements
        data["inline_results"] = []
        data["tables"] = _scan_tables(ws2, area["table_start_row"]) if ws2 is not None else []

    if data["area_code"]:
        known_area = area_registry.get_area(data["area_code"])
        data["area_name"] = known_area["area_name"] if known_area else data["area_code"].replace("_", " ").title()
    return data
